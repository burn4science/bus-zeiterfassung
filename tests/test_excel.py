from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bus_zeiterfassung.models import TimeEntry
from bus_zeiterfassung.services.excel import fill_template


def test_fill_template_writes_header(tmp_export_dir: Path) -> None:
    out = fill_template([], 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws["A1"].value == "Dienstzeitblatt für Monat: April"
    assert ws["G1"].value == "Jahr:"
    assert ws["H1"].value == 2026


def test_fill_template_writes_session_cells(tmp_export_dir: Path) -> None:
    entries = [
        TimeEntry(day=date(2026, 4, 1), start=time(8, 0), end=time(12, 30)),
        TimeEntry(day=date(2026, 4, 2), start=time(9, 0), end=time(11, 0)),
    ]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None

    # Day 1 → row 7, day 2 → row 8. Session 1 goes into B/C.
    assert ws["B7"].value == time(8, 0)
    assert ws["C7"].value == time(12, 30)
    assert ws["B8"].value == time(9, 0)
    assert ws["C8"].value == time(11, 0)


def test_fill_template_single_session_daily_total(tmp_export_dir: Path) -> None:
    entries = [
        TimeEntry(day=date(2026, 4, 1), start=time(8, 0), end=time(12, 30)),   # 4h30m
        TimeEntry(day=date(2026, 4, 2), start=time(9, 0), end=time(11, 0)),    # 2h
    ]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None

    j7 = ws["J7"].value
    assert isinstance(j7, time), f"J7 should be a time value, got {type(j7)}: {j7!r}"
    assert j7.hour == 4 and j7.minute == 30, f"J7 expected 04:30, got {j7}"

    j8 = ws["J8"].value
    assert isinstance(j8, time), f"J8 should be a time value, got {type(j8)}: {j8!r}"
    assert j8.hour == 2 and j8.minute == 0, f"J8 expected 02:00, got {j8}"


def test_fill_template_multi_session_daily_total(tmp_export_dir: Path) -> None:
    """4 sessions on one day must sum to exactly 67 minutes (01:07)."""
    entries = [
        TimeEntry(day=date(2026, 4, 19), start=time(14, 30), end=time(14, 35)),  # 5 min
        TimeEntry(day=date(2026, 4, 19), start=time(14, 38), end=time(14, 39)),  # 1 min
        TimeEntry(day=date(2026, 4, 19), start=time(15, 0),  end=time(16, 0)),   # 60 min
        TimeEntry(day=date(2026, 4, 19), start=time(17, 11), end=time(17, 12)),  # 1 min
    ]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None

    # Verify all 4 session pairs were written (day 19 → row 25)
    assert ws["B25"].value == time(14, 30)
    assert ws["C25"].value == time(14, 35)
    assert ws["D25"].value == time(14, 38)
    assert ws["E25"].value == time(14, 39)
    assert ws["F25"].value == time(15, 0)
    assert ws["G25"].value == time(16, 0)
    assert ws["H25"].value == time(17, 11)
    assert ws["I25"].value == time(17, 12)

    j25 = ws["J25"].value
    assert isinstance(j25, time), f"J25 should be a time value, got {type(j25)}: {j25!r}"
    assert j25.hour == 1 and j25.minute == 7, f"J25 expected 01:07, got {j25}"


def test_fill_template_k_column_formula_preserved(tmp_export_dir: Path) -> None:
    entries = [TimeEntry(day=date(2026, 4, 1), start=time(8, 0), end=time(9, 0))]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # K column formulas must remain untouched
    assert isinstance(ws["K7"].value, str) and ws["K7"].value.startswith("=")


def test_fill_template_j38_formula_preserved(tmp_export_dir: Path) -> None:
    out = fill_template([], 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # J38 monthly total formula must remain untouched
    assert isinstance(ws["J38"].value, str) and ws["J38"].value.startswith("=")


def test_fill_template_skips_entries_outside_month(tmp_export_dir: Path) -> None:
    entries = [
        TimeEntry(day=date(2026, 3, 31), start=time(8, 0), end=time(9, 0)),
        TimeEntry(day=date(2026, 4, 1), start=time(8, 0), end=time(9, 0)),
        TimeEntry(day=date(2026, 5, 1), start=time(8, 0), end=time(9, 0)),
    ]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws["B7"].value == time(8, 0)   # April 1 written
    assert ws["B37"].value is None        # March 31 filtered out


def test_fill_template_ignores_entries_without_start(tmp_export_dir: Path) -> None:
    entries = [TimeEntry(day=date(2026, 4, 1), start=None, end=None)]
    out = fill_template(entries, 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws["B7"].value is None


def test_fill_template_writes_user_name(tmp_export_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from bus_zeiterfassung import config

    monkeypatch.setattr(config.settings, "employee_name", "Max Mustermann")
    out = fill_template([], 2026, 4)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws["A3"].value == "Name:"
    assert ws["B3"].value == "Max Mustermann"


@pytest.mark.parametrize(
    ("month", "name"),
    [(1, "Januar"), (3, "März"), (7, "Juli"), (12, "Dezember")],
)
def test_german_month_in_header(tmp_export_dir: Path, month: int, name: str) -> None:
    out = fill_template([], 2026, month)
    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws["A1"].value == f"Dienstzeitblatt für Monat: {name}"
