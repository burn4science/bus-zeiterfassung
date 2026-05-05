import logging
import ssl
import tempfile
import urllib.request
from collections import defaultdict
from collections.abc import Generator, Sequence
from contextlib import contextmanager
from datetime import date as _date
from datetime import datetime as _dt
from datetime import time as _time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from bus_zeiterfassung.config import settings
from bus_zeiterfassung.models import TimeEntry

logger = logging.getLogger(__name__)

GERMAN_MONTHS = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember",
}

SESSION_COLS = (("B", "C"), ("D", "E"), ("F", "G"), ("H", "I"))

_THIN = Side(style="thin")
_GRID = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SIGNATURE_HEIGHT_PX = 31


def _apply_table_borders(ws: Worksheet) -> None:
    for row in ws.iter_rows(min_row=5, max_row=37, min_col=1, max_col=11):
        for cell in row:
            cell.border = _GRID


def _fetch_signature_to_tempfile(url: str) -> Path:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(url, context=ctx) as resp:
        data = resp.read()
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(data)
    return Path(tmp.name)


@contextmanager
def _signature_image() -> Generator[Path | None, None, None]:
    if settings.signature_path and settings.signature_path.exists():
        logger.info("signature: using local file %s", settings.signature_path)
        yield settings.signature_path
    elif settings.signature_url:
        logger.info("signature: fetching from %s", settings.signature_url)
        tmp = _fetch_signature_to_tempfile(settings.signature_url)
        try:
            yield tmp
        finally:
            tmp.unlink(missing_ok=True)
    else:
        yield None


def _insert_signature(ws: Worksheet, path: Path, anchor: str) -> None:
    img = XlImage(str(path))
    img.width = round(img.width * _SIGNATURE_HEIGHT_PX / img.height)
    img.height = _SIGNATURE_HEIGHT_PX
    img.anchor = anchor
    ws.add_image(img)


# --- Data helpers ---


def _session_duration_minutes(start: _time, end: _time) -> int:
    return (_dt.combine(_date.min, end) - _dt.combine(_date.min, start)).seconds // 60


def _group_entries_by_day(
    entries: Sequence[TimeEntry], year: int, month: int
) -> dict[int, list[TimeEntry]]:
    by_day: dict[int, list[TimeEntry]] = defaultdict(list)
    for e in entries:
        if e.day.year == year and e.day.month == month and e.start is not None:
            by_day[e.day.day].append(e)
    return by_day


# --- Worksheet-writing helpers ---


def _write_header(ws: Worksheet, year: int, month: int) -> None:
    ws["A1"] = f"Dienstzeitblatt für Monat: {GERMAN_MONTHS[month]}"
    ws["G1"] = "Jahr:"
    ws["H1"] = year
    ws["A3"] = "Name:"
    ws["B3"] = settings.employee_name


def _write_day_row(ws: Worksheet, day: int, sessions: list[TimeEntry]) -> None:
    row = day + 6
    for idx, session in enumerate(sessions[:4]):
        start_col, end_col = SESSION_COLS[idx]
        ws[f"{start_col}{row}"] = session.start
        if session.end is not None:
            ws[f"{end_col}{row}"] = session.end

    total_min = sum(
        _session_duration_minutes(s.start, s.end)
        for s in sessions
        if s.start is not None and s.end is not None
    )
    if total_min > 0:
        h, m = divmod(total_min, 60)
        # 1s bias so LibreOffice's hh:mm floor never rounds down
        ws[f"J{row}"] = _time(h, m, 1)


def _write_session_data(ws: Worksheet, by_day: dict[int, list[TimeEntry]]) -> None:
    for day, sessions in by_day.items():
        sessions.sort(key=lambda e: e.start)  # type: ignore[arg-type,return-value]
        _write_day_row(ws, day, sessions)


# --- Formatting helpers ---


def _apply_number_formats(ws: Worksheet) -> None:
    for col in "BCDEFGHIJ":
        for row_num in range(7, 38):
            ws[f"{col}{row_num}"].number_format = "hh:mm"
    ws["J38"].number_format = "hh:mm"
    ws["J38"].font = Font(size=11, bold=True)


def _apply_fonts(ws: Worksheet) -> None:
    font_bands = (
        (1, 5, 11),
        (5, 7, 10),
        (7, 38, 11),
        (38, ws.max_row + 1, 10),
    )
    for start, stop, size in font_bands:
        font = Font(name="Calibri", size=size)
        for row_num in range(start, stop):
            for cell in ws[row_num]:
                cell.font = font


def _apply_column_widths(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 6
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 7
    ws.column_dimensions["K"].width = 20


def _configure_page_setup(ws: Worksheet) -> None:
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "portrait"


# --- Output helper ---


def _build_output_path(year: int, month: int) -> Path:
    settings.export_dir.mkdir(parents=True, exist_ok=True)
    stem = f"Dienstzeitblatt {settings.employee_name} {GERMAN_MONTHS[month]}"
    return settings.export_dir / f"{stem}.xlsx"


# --- Public API ---


def fill_template(entries: Sequence[TimeEntry], year: int, month: int) -> Path:
    """Fill the Dienstzeitblatt template for a given month and write it to EXPORT_DIR.

    Cell layout is documented in docs/template-mapping.md. J7:J37 formulas are
    overwritten to sum all 4 sessions. K formulas are never touched.
    """
    wb = load_workbook(settings.template_path, keep_vba=False)
    ws = wb.active
    assert ws is not None

    _write_header(ws, year, month)
    _write_session_data(ws, _group_entries_by_day(entries, year, month))
    _apply_number_formats(ws)
    _apply_table_borders(ws)
    _apply_fonts(ws)
    _apply_column_widths(ws)
    _configure_page_setup(ws)

    out_path = _build_output_path(year, month)
    with _signature_image() as sig:
        if sig:
            _insert_signature(ws, sig, anchor="J43")
        wb.save(out_path)

    return out_path
